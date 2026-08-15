import io
import json
import os
import calendar
import re
import hashlib
from datetime import datetime, date

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from roster_data import (
    employees as DEFAULT_EMPLOYEES,
    days,
    hours,
    preferences as DEFAULT_PREFERENCES,
    STAFFING_NEEDS,
)
from rules_engine import check_rules
from scheduler import generate_month_schedule

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PREFERENCES_FILE = os.path.join(BASE_DIR, "preferences.json")
EMPLOYEES_FILE = os.path.join(BASE_DIR, "employees.xlsx")


# --- PREFERENCE PERSISTENCE ---

def load_preferences():
    """Load day-off preferences from file, or seed from defaults."""
    if os.path.exists(PREFERENCES_FILE):
        with open(PREFERENCES_FILE, "r", encoding="utf-8") as f:
            loaded = {}
            for key, value in json.load(f).items():
                key_text = str(key).strip()
                loaded[key_text] = value
                if key_text.isdigit():
                    loaded[int(key_text)] = value
            return loaded
    return DEFAULT_PREFERENCES.copy()


def save_preferences(prefs):
    """Save day-off preferences to a JSON file."""
    with open(PREFERENCES_FILE, "w", encoding="utf-8") as f:
        json.dump(prefs, f, indent=2)


# --- EMPLOYEE LOADING FROM EXCEL ---

def _parse_day_off_cell(value):
    """
    Parse a day-off Excel cell.

    Returns None (whole day off), a list of hour strings, or an empty list.
    """
    if value is None:
        return []
    value = str(value).strip()
    if value == "" or value.lower() == "nan":
        return []
    if value.upper() == "OFF":
        return None
    if value.upper() in {"PREFERRED", "PREF", "P"}:
        return None
    # Accept individual hours and inclusive ranges, e.g. 08,09 or 08-10.
    parts = [p.strip() for p in value.replace(";", ",").replace(" ", ",").split(",") if p.strip()]
    parsed = []
    for part in parts:
        if "-" in part:
            start, end = [x.strip() for x in part.split("-", 1)]
            if start.isdigit() and end.isdigit():
                parsed.extend(f"{h:02d}" for h in range(int(start), int(end) + 1) if 0 <= h <= 23)
        elif part.isdigit() and 0 <= int(part) <= 23:
            parsed.append(f"{int(part):02d}")
    return list(dict.fromkeys(parsed))


def _payroll_role_and_cross_train(job_text):
    """Map the hotel's payroll job labels to roster departments."""
    job = str(job_text or "").strip().upper()
    if "HK" in job:
        role = "HK"
    elif "FB" in job or "OP" in job:
        role = "FB"
    elif "AC" in job:
        role = "AC"
    elif "GSA" in job or "RSVN" in job or "FO" in job:
        role = "FO"
    elif "AT" in job or "DM" in job:
        role = "AT"
    elif "EN" in job:
        role = "EN"
    else:
        role = "EN"

    cross_train = []
    if "FB" in job and role == "FO":
        cross_train.append("FB")
    if "DM" in job and role in {"FO", "EN"}:
        cross_train.extend(["AT", "EN"] if role == "FO" else ["AT"])
    if "AT" in job and role != "AT":
        cross_train.append("AT")
    if role == "AT":
        cross_train.append("EN")
        if "FO" in job or "GSA" in job:
            cross_train.append("FO")
        if "FB" in job:
            cross_train.append("FB")
    if role == "AC":
        cross_train.extend(["HR", "FO", "FB"])
    if role == "EN":
        cross_train.append("HK")
    if role == "HK":
        cross_train.append("FB")
    return role, list(dict.fromkeys(cross_train))


def _load_payroll_style_excel(file_path):
    """Read the hotel's payroll roster layout (JOB/Name rows and date columns)."""
    raw = pd.read_excel(file_path, header=None)
    employees = []
    seen_ids = set()
    for row_index, row in raw.iterrows():
        # The payroll sheet has four header rows followed by employee rows.
        if row_index < 6:
            continue
        code = "" if pd.isna(row.iloc[1]) else str(row.iloc[1]).strip()
        job = "" if pd.isna(row.iloc[2]) else str(row.iloc[2]).strip()
        name = "" if pd.isna(row.iloc[3]) else str(row.iloc[3]).strip()
        if not name or not job or name.lower() == "nan" or job.lower() == "nan":
            continue
        upper_job = job.upper()
        if upper_job.startswith("CHECK") or upper_job.startswith("FO "):
            continue
        # Summary rows have no person-like name or use aggregate labels.
        if any(token in upper_job for token in ("MAIN TEAM", "PT+S", "TN (", "ALL (", "SL (")):
            continue
        emp_id = code or f"PAY-{row_index + 1:04d}"
        if emp_id in seen_ids:
            emp_id = f"{emp_id}-{row_index + 1}"
        seen_ids.add(emp_id)
        role, cross_train = _payroll_role_and_cross_train(job)
        employees.append({"id": emp_id, "name": name, "role": role, "cross_train": cross_train})
    if not employees:
        raise ValueError("No employee rows found in payroll workbook")
    return employees, {}


def load_employees_from_excel(file_path):
    """Read employees and preferences from the Excel template."""
    df = pd.read_excel(file_path)
    required = {"id", "name", "role", "cross_train"}
    missing = required - set(df.columns)
    if missing:
        # The hotel also uses a payroll workbook with JOB/Name columns and
        # dates across the sheet. Read that layout instead of falling back to
        # demo employees when it is uploaded.
        if not required.issubset(set(df.columns)):
            return _load_payroll_style_excel(file_path)

    employees = []
    preferences = {}

    for _, row in df.iterrows():
        if pd.isna(row["id"]) or pd.isna(row["name"]) or pd.isna(row["role"]):
            raise ValueError("Every employee row must include id, name, and role")
        raw_id = row["id"]
        if isinstance(raw_id, float) and raw_id.is_integer():
            emp_id = int(raw_id)
        else:
            emp_id = str(raw_id).strip()
        role = str(row["role"]).strip()
        if role.lower() in {"security", "sec"}:
            role = "SEC"
        cross_raw = row.get("cross_train", "")
        if pd.isna(cross_raw):
            cross_raw = ""
        cross_train = [c.strip() for c in str(cross_raw).split(",") if c.strip()]
        cross_train = ["SEC" if c.lower() in {"security", "sec"} else c for c in cross_train]

        employees.append({
            "id": emp_id,
            "name": str(row["name"]).strip(),
            "role": role,
            "cross_train": cross_train,
        })

        emp_prefs = {}
        # Support the original weekly format (Mon-Sun) and monthly date columns (01 Aug, 02 Aug...).
        preference_columns = [(day, day) for day in days if day in df.columns]
        for column in df.columns:
            if column in {"id", "name", "role", "cross_train", *days}:
                continue
            try:
                header_text = re.sub(r"\s+", " ", str(column).replace("\n", " ")).strip()
                header_parts = header_text.split()
                weekday_names = {"Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"}
                if header_parts and header_parts[0] in weekday_names:
                    date_text = " ".join(header_parts[-2:])
                    parsed_date = datetime.strptime(f"{date_text} {datetime.now().year}", "%d %b %Y")
                    preference_columns.append((column, parsed_date.strftime("%Y-%m-%d")))
                else:
                    date_text = " ".join(header_parts[-2:])
                    parsed_date = datetime.strptime(f"{date_text} {datetime.now().year}", "%d %b %Y")
                    preference_columns.append((column, parsed_date.strftime("%Y-%m-%d")))
            except ValueError:
                continue
        for column, day in preference_columns:
            cell = row.get(column, "")
            parsed = _parse_day_off_cell(cell)
            if parsed is None:
                emp_prefs[day] = None  # whole day off
            elif parsed:
                emp_prefs[day] = parsed

        if emp_prefs:
            preferences[emp_id] = emp_prefs

    return employees, preferences


def get_employees_and_preferences():
    """Return the active employee list and preferences."""
    if os.path.exists(EMPLOYEES_FILE):
        try:
            return load_employees_from_excel(EMPLOYEES_FILE)
        except Exception as e:
            st.error(f"Could not read {EMPLOYEES_FILE}: {e}. Falling back to defaults.")
    return DEFAULT_EMPLOYEES.copy(), DEFAULT_PREFERENCES.copy()


# --- SCHEDULE HELPERS ---

def empty_schedule():
    return {day: {hour: [] for hour in hours} for day in days}


def staffing_gaps(schedule):
    gaps = []
    for day in schedule.keys():
        for hour in hours:
            assigned = set(schedule.get(day, {}).get(hour, []))
            for role, needs in STAFFING_NEEDS.items():
                required = needs.get(hour, 0)
                eligible = [e for e in employees if e["id"] in assigned and (e["role"] == role or role in e.get("cross_train", []) or (role == "HR" and e["role"] in {"AC", "MD"}))]
                if len(eligible) < required:
                    gaps.append({"Day": day, "Hour": f"{hour}:00", "Department": role, "Required": required, "Assigned": len(eligible), "Shortage": required - len(eligible)})
    return pd.DataFrame(gaps)


def staffing_plan_table(month_start):
    """Compact editable planning table using only meaningful coverage start hours."""
    month_start = month_start.replace(day=1)
    departments = [d for d in STAFFING_NEEDS if d != "SEC"]
    active_hours = []
    for index, hour in enumerate(hours):
        previous = hours[index - 1] if index else None
        if any(
            STAFFING_NEEDS[d].get(hour, 0) > 0
            and (previous is None or STAFFING_NEEDS[d].get(previous, 0) != STAFFING_NEEDS[d].get(hour, 0))
            for d in departments
        ):
            active_hours.append(hour)
    rows = []
    for day_number in range(1, calendar.monthrange(month_start.year, month_start.month)[1] + 1):
        current = date(month_start.year, month_start.month, day_number)
        for hour in active_hours:
            row = {"Date": current.strftime("%Y-%m-%d"), "Hour": f"{hour}:00"}
            row.update({dept: STAFFING_NEEDS[dept].get(hour, 0) for dept in departments})
            rows.append(row)
    return pd.DataFrame(rows)


def _format_hours(emp_hours, view_hours):
    """Format an employee's assigned hours for display, filtered by view_hours."""
    visible = sorted(h for h in emp_hours if h in view_hours)
    if not visible:
        return "—"
    nums = sorted(int(h) for h in visible)
    # Show a clear arrival-to-departure window for employees.
    if len(nums) == 1:
        return f"{nums[0]:02d}"
    return f"{nums[0]:02d}-{nums[-1]:02d}"


def _parse_work_window(value):
    """Parse an editable schedule cell such as OFF, 08, or 08-16."""
    text = str(value or "").strip().upper()
    if not text or text == "OFF":
        return []
    first = text.split(",")[0].strip()
    if "-" in first:
        start, end = [part.strip() for part in first.split("-", 1)]
        if start.isdigit() and end.isdigit():
            return [f"{h:02d}" for h in range(int(start), int(end) + 1) if 0 <= h <= 23]
    if first.isdigit() and 0 <= int(first) <= 23:
        return [f"{int(first):02d}"]
    return []


def apply_monthly_schedule_edits(schedule, employees, roster_df, visible_df, edited_rows, working_month, preferences):
    """Apply only the cells changed in the monthly editor to the live roster."""
    employee_by_name = {e["name"]: e for e in employees}
    visible_rows = visible_df.reset_index(drop=True)
    changed = 0
    for row_position, changes in (edited_rows or {}).items():
        try:
            row = visible_rows.iloc[int(row_position)]
        except (ValueError, TypeError, IndexError):
            continue
        emp = employee_by_name.get(row.get("Employee"))
        if not emp:
            continue
        original_match = roster_df[roster_df["Employee"] == emp["name"]]
        if original_match.empty:
            continue
        original_row = original_match.iloc[0]
        for column, raw_value in changes.items():
            if column not in roster_df.columns[2:]:
                continue
            new_value = str(raw_value or "").strip().upper()
            old_value = str(original_row[column]).strip().upper()
            if new_value == old_value:
                continue
            new_hours = _parse_work_window(new_value)
            if new_value and new_value != "OFF" and not new_hours:
                continue
            day_key = datetime.strptime(
                f"{column} {working_month.year}", "%d %b %Y"
            ).strftime("%Y-%m-%d")
            if day_key in schedule:
                for hour in hours:
                    schedule[day_key][hour] = [
                        eid for eid in schedule[day_key][hour] if eid != emp["id"]
                    ]
                for hour in new_hours:
                    if emp["id"] not in schedule[day_key][hour]:
                        schedule[day_key][hour].append(emp["id"])
            if not new_value or new_value == "OFF":
                preferences.setdefault(emp["id"], {})[day_key] = None
            else:
                preferences.setdefault(emp["id"], {}).pop(day_key, None)
            changed += 1
    return changed


def schedule_to_employee_view(schedule, employees, view_hours):
    """Employee-centric roster: rows=employees, columns=days, cells=hours."""
    # Build reverse lookup: emp_id -> {day: [hours]}
    emp_assignments = {e["id"]: {d: [] for d in days} for e in employees}
    for day in days:
        for hour in hours:
            for eid in schedule.get(day, {}).get(hour, []):
                if eid in emp_assignments:
                    emp_assignments[eid][day].append(hour)

    data = []
    for emp in sorted(employees, key=lambda e: (e["role"], e["name"])):
        row = {
            "Employee": emp["name"],
            "Dept": emp["role"],
        }
        for day in days:
            assigned = emp_assignments[emp["id"]].get(day, [])
            row[day] = _format_hours(assigned, view_hours) if assigned else "OFF"
        data.append(row)
    return pd.DataFrame(data)


def schedule_to_month_employee_view(schedule, employees, month_start, view_hours):
    """Employee-centric monthly roster with one column for each calendar date."""
    month_start = month_start.replace(day=1)
    num_days = calendar.monthrange(month_start.year, month_start.month)[1]
    month_dates = [date(month_start.year, month_start.month, d) for d in range(1, num_days + 1)]
    date_schedule = all(len(str(k)) == 10 and str(k)[4] == "-" for k in schedule.keys())
    if date_schedule:
        by_id = {e["id"]: e for e in employees}
        assignments = {e["id"]: {d: [] for d in month_dates} for e in employees}
        for current in month_dates:
            key = current.strftime("%Y-%m-%d")
            for hour in hours:
                for eid in schedule.get(key, {}).get(hour, []):
                    if eid in assignments:
                        assignments[eid][current].append(hour)
        rows = []
        for emp in sorted(employees, key=lambda e: (e["role"], e["name"])):
            row = {"Employee": emp["name"], "Dept": emp["role"]}
            for current in month_dates:
                row[current.strftime("%d %b")] = _format_hours(assignments[emp["id"]][current], view_hours) if assignments[emp["id"]][current] else "OFF"
            rows.append(row)
        return pd.DataFrame(rows)
    weekly = schedule_to_employee_view(schedule, employees, view_hours)
    rows = []
    for _, source in weekly.iterrows():
        row = {"Employee": source["Employee"], "Dept": source["Dept"]}
        for current in month_dates:
            day_name = current.strftime("%a")
            row[current.strftime("%d %b")] = source.get(day_name, "OFF")
        rows.append(row)
    return pd.DataFrame(rows)


def schedule_to_department_time_view(schedule, employees, selected_day, view_hours):
    """Daily coverage view: rows are times, columns are departments."""
    by_id = {e["id"]: e for e in employees}
    departments = sorted(STAFFING_NEEDS.keys())
    rows = []
    for hour in view_hours:
        row = {"Time": f"{hour}:00"}
        assigned = schedule.get(selected_day, {}).get(hour, [])
        for dept in departments:
            names = []
            for eid in assigned:
                emp = by_id.get(eid)
                if emp and (emp["role"] == dept or dept in emp.get("cross_train", []) or (dept == "HR" and emp["role"] in {"AC", "MD"})):
                    names.append(emp["name"])
            row[dept] = ", ".join(names) if names else "—"
        rows.append(row)
    return pd.DataFrame(rows)


def schedule_to_month_hours_summary(schedule, employees, month_start):
    """Monthly hours summary: rows are employees, columns are calendar dates."""
    month_start = month_start.replace(day=1)
    num_days = calendar.monthrange(month_start.year, month_start.month)[1]
    dates = [date(month_start.year, month_start.month, d) for d in range(1, num_days + 1)]
    day_schedule_is_monthly = all(len(str(k)) == 10 and str(k)[4] == "-" for k in schedule.keys())
    rows = []
    for emp in sorted(employees, key=lambda e: (e["role"], e["name"])):
        row = {"Employee": emp["name"], "Dept": emp["role"]}
        for current in dates:
            day_name = current.strftime("%Y-%m-%d") if day_schedule_is_monthly else current.strftime("%a")
            assigned = schedule.get(day_name, {})
            row[current.strftime("%d %b")] = sum(emp["id"] in ids for ids in assigned.values())
        row["Monthly Total"] = sum(v for k, v in row.items() if k not in {"Employee", "Dept"})
        rows.append(row)
    return pd.DataFrame(rows)


def style_month_table(df, month_start):
    """Add compact calendar-style week separators using the approved blue/white palette."""
    month_start = month_start.replace(day=1)
    num_days = calendar.monthrange(month_start.year, month_start.month)[1]
    styles = [{"selector": "th", "props": [("background-color", "#0E5294"), ("color", "#FFFFFF"), ("font-weight", "600"), ("white-space", "nowrap")]},
              {"selector": "td", "props": [("white-space", "nowrap")] }]
    for day_number in range(1, num_days + 1):
        current = date(month_start.year, month_start.month, day_number)
        if current.weekday() == 6:
            column = current.strftime("%d %b")
            if column in df.columns:
                styles.append({"selector": f"th.col{df.columns.get_loc(column)}", "props": [("border-left", "3px solid #0E5294")]})
                styles.append({"selector": f"td.col{df.columns.get_loc(column)}", "props": [("border-left", "3px solid #0E5294")]})
    return df.style.set_table_styles(styles)


# --- PAGE SETUP ---

st.set_page_config(page_title="Hotel Smart Roster Manager", layout="wide")
st.markdown("""
<style>
:root { --blue:#0E5294; --white:#ffffff; --black:#000000; }
.stApp { background: var(--white); color: var(--black); }
[data-testid="stSidebar"] { background: var(--blue); }
[data-testid="stSidebar"] * { color: #ffffff !important; }
[data-testid="stSidebar"] input, [data-testid="stSidebar"] textarea,
[data-testid="stSidebar"] [data-baseweb="select"] > div { background:#ffffff !important; color:#000000 !important; }
[data-testid="stSidebar"] [data-baseweb="select"] span,
[data-testid="stSidebar"] input { color:#000000 !important; }
[data-testid="stSidebar"] button { background:#0E5294 !important; color:#ffffff !important; border:1px solid #ffffff !important; }
[data-testid="stSidebar"] button p { color:#ffffff !important; }
[data-baseweb="tag"] { background:#0E5294 !important; color:#ffffff !important; padding:4px 8px !important; margin:2px 4px 2px 0 !important; max-width:none !important; overflow:visible !important; }
[data-baseweb="tag"] span { color:#ffffff !important; display:inline-block !important; max-width:none !important; overflow:visible !important; }
[data-baseweb="tag"] svg { fill:#ffffff !important; color:#ffffff !important; }
[data-testid="stSidebar"] [data-baseweb="select"] { background:#ffffff !important; }
[data-testid="stSidebar"] [data-baseweb="select"] > div { overflow:visible !important; min-height:42px !important; }
[data-testid="stSidebar"] [data-baseweb="select"] input { background:#ffffff !important; color:#000000 !important; }
[data-testid="stAlert"] { background:#ffffff !important; color:#000000 !important; border:1px solid #0E5294 !important; border-radius:8px !important; box-shadow:none !important; }
[data-testid="stAlert"] * { color:#000000 !important; }
[data-testid="stFileUploader"] section { background:#ffffff !important; border:1px solid #0E5294 !important; }
[data-testid="stFileUploader"] section > div,
[data-testid="stFileUploader"] small,
[data-testid="stFileUploader"] span,
[data-testid="stFileUploader"] p { color:#000000 !important; }
[data-testid="stFileUploader"] button { background:#0E5294 !important; color:#ffffff !important; border:1px solid #0E5294 !important; }
[data-testid="stFileUploader"] button p { color:#ffffff !important; }
[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] { background:#ffffff !important; border:1px solid #0E5294 !important; }
[data-testid="stAlert"] { background:#ffffff !important; color:#000000 !important; border:1px solid #0E5294 !important; }
h1,h2,h3 { color: var(--blue); font-family: Georgia, serif; }
p, label, [data-testid="stMarkdownContainer"] { color: #000000; }
[data-testid="stCaptionContainer"] { color:#000000 !important; }
[data-testid="stMetric"] { background:#ffffff; border:1px solid #0E5294; border-radius:10px; padding:12px; box-shadow:none; }
.nyonya-banner { background:#0E5294; color:#ffffff; padding:22px 28px; border-radius:16px; margin-bottom:18px; border-left:8px solid #ffffff; }
.nyonya-banner h1 { color:#ffffff; margin:0; }
.nyonya-banner p { margin:.35rem 0 0; color:#ffffff; }
</style>
<div class="nyonya-banner"><h1>Little Nyonya Roster House</h1><p>Hotel operations dashboard · Phuket City · Sino-colonial hospitality</p></div>
""", unsafe_allow_html=True)

# Load current data
employees, file_preferences = get_employees_and_preferences()
current_preferences = load_preferences()

# Merge: uploaded Excel preferences seed the roster; saved sidebar edits take priority.
for eid, prefs in file_preferences.items():
    current_preferences[eid] = prefs
saved_preferences = load_preferences()
for eid, prefs in saved_preferences.items():
    current_preferences[eid] = prefs

# Initialize Session State
if "schedule" not in st.session_state:
    st.session_state["schedule_month"] = date.today().replace(day=1)
    st.session_state["working_date"] = st.session_state["schedule_month"]
    st.session_state["schedule"] = generate_month_schedule(
        employees, st.session_state["schedule_month"], current_preferences, STAFFING_NEEDS
    )

# --- SIDEBAR ---

st.sidebar.header("Controls")

# 1. Excel upload
st.sidebar.markdown("### Step 1 — Upload employees")
uploaded_file = st.sidebar.file_uploader(
    "Upload employee Excel",
    type=["xlsx"],
    help="Columns: id, name, role, cross_train, Mon, Tue, Wed, Thu, Fri, Sat, Sun. Use OFF for whole day off, or hours/ranges like 08-10,22.",
)

if uploaded_file is not None:
    upload_bytes = uploaded_file.getvalue()
    upload_signature = hashlib.sha256(upload_bytes).hexdigest()
    # UploadedFile remains populated after st.rerun(). Only process a file once
    # per content signature, otherwise the page can enter an endless rerun loop.
    if st.session_state.get("last_uploaded_signature") != upload_signature:
        with open(EMPLOYEES_FILE, "wb") as f:
            f.write(upload_bytes)
        # Uploaded monthly preferences start a fresh roster cycle; discard stale
        # manager edits from a previous employee file.
        if os.path.exists(PREFERENCES_FILE):
            os.remove(PREFERENCES_FILE)
        # Rebuild the draft using the newly uploaded employee IDs and names.
        st.session_state.pop("schedule", None)
        st.session_state["last_uploaded_signature"] = upload_signature
        st.session_state["upload_complete"] = True
        st.sidebar.success(f"Saved {EMPLOYEES_FILE}. Refreshing...")
        st.rerun()

if st.sidebar.button("Download blank employee template"):
    template = pd.DataFrame(
        [{"id": "", "name": "", "role": "", "cross_train": ""} | {d: "" for d in days}]
    )
    buffer = io.BytesIO()
    template.to_excel(buffer, index=False, engine="openpyxl")
    st.sidebar.download_button(
        label="Download template",
        data=buffer.getvalue(),
        file_name="employee_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if st.sidebar.button("Download day-off preference file"):
    preference_rows = []
    for employee in employees:
        preference_rows.append({
            "id": employee["id"],
            "name": employee["name"],
            "role": employee["role"],
            "cross_train": ", ".join(employee.get("cross_train", [])),
            **{day: "" for day in days},
        })
    preference_template = pd.DataFrame(preference_rows)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        preference_template.to_excel(writer, index=False, sheet_name="Day-Off Preferences")
        notes = pd.DataFrame({"Instructions": [
            "Leave a day blank when the employee is available.",
            "Enter OFF for a full day off.",
            "Enter hour ranges such as 08-10,22 for specific hours off.",
            "Upload this file using Upload employee Excel after editing.",
        ]})
        notes.to_excel(writer, index=False, sheet_name="Instructions")
    st.sidebar.download_button(
        label="Download preferences Excel",
        data=buffer.getvalue(),
        file_name="employee_day_off_preferences.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.sidebar.divider()

# 2. Schedule actions
st.sidebar.markdown("### Step 2 — Generate first draft")
if st.sidebar.button("Generate Rule-Based Schedule"):
    target_month = st.session_state.get("working_date", date.today().replace(day=1))
    st.session_state["schedule_month"] = target_month.replace(day=1)
    st.session_state["schedule"] = generate_month_schedule(
        employees, st.session_state["schedule_month"], current_preferences, STAFFING_NEEDS
    )
    st.session_state["generation_complete"] = True
    st.rerun()

if st.sidebar.button("Clear Schedule"):
    st.session_state["schedule"] = empty_schedule()
    st.rerun()

if st.sidebar.button("Save Current Roster"):
    with open(os.path.join(BASE_DIR, "saved_roster.json"), "w", encoding="utf-8") as f:
        json.dump(st.session_state["schedule"], f, indent=2)
    st.sidebar.success("Roster saved.")

if st.sidebar.button("Load Saved Roster"):
    saved_path = os.path.join(BASE_DIR, "saved_roster.json")
    if os.path.exists(saved_path):
        with open(saved_path, "r", encoding="utf-8") as f:
            st.session_state["schedule"] = json.load(f)
        st.sidebar.success("Roster loaded.")
        st.rerun()

st.sidebar.divider()

if False:
    _disabled_sidebar_code = """
# 3. Day-off preference editor
st.sidebar.markdown("### Step 3 — Adjust preferences")
st.sidebar.caption("The employee list below comes from the file uploaded in Step 1.")
st.sidebar.header("Staff Day-Off Preferences")

selected_pref_emp = st.sidebar.selectbox(
    "Select Employee",
    employees,
    format_func=lambda e: f"{e['id']}: {e['name']} ({e['role']})",
)

emp_prefs = current_preferences.get(selected_pref_emp["id"], {})
monthly_pref_dates = sorted(d for d, v in emp_prefs.items() if len(str(d)) == 10 and v is None)
preference_options = days + monthly_pref_dates
whole_days_off = [d for d, v in emp_prefs.items() if v is None and d in preference_options]

selected_days_off = st.sidebar.multiselect(
    "Whole days off",
    preference_options,
    default=whole_days_off,
)

if st.sidebar.button("Save Preferences"):
    new_prefs = {k: v for k, v in emp_prefs.items() if k not in preference_options}
    for day in preference_options:
        if day in selected_days_off:
            new_prefs[day] = None
    current_preferences[selected_pref_emp["id"]] = new_prefs
    save_preferences(current_preferences)
    st.sidebar.success(f"Saved {selected_pref_emp['name']}'s preferences!")
    st.rerun()

if st.sidebar.button("Reset to Defaults"):
    if os.path.exists(PREFERENCES_FILE):
        os.remove(PREFERENCES_FILE)
    current_preferences = DEFAULT_PREFERENCES.copy()
    st.sidebar.success("Preferences reset to defaults.")
    st.rerun()

"""
# --- MAIN DISPLAY ---

if st.session_state.pop("upload_complete", False):
    st.success("Employee file uploaded and loaded. The previous draft was cleared; generate a new draft in Step 2.")
if st.session_state.pop("generation_complete", False):
    st.success("Draft roster generated and tables updated.")

# 1. Rules Status
errors = check_rules(st.session_state["schedule"], employees, current_preferences, hours)
gaps_df = staffing_gaps(st.session_state["schedule"])
total_assigned = sum(len(v) for d in st.session_state["schedule"].values() for v in d.values())
metric_cols = st.columns(4)
metric_cols[0].metric("Employees", len(employees))
metric_cols[1].metric("Assigned hours", total_assigned)
metric_cols[2].metric("Rule conflicts", len(errors))
metric_cols[3].metric("Staffing gaps", len(gaps_df))
if errors:
    st.error("Schedule Conflicts Detected:")
    for err in errors:
        st.write(err)
elif gaps_df.empty:
    st.success("Schedule looks valid so far!")
else:
    st.warning("No rule conflicts, but the staffing requirements are not fully covered yet. Review the staffing gaps before publishing.")

if not gaps_df.empty:
    st.warning(f"{len(gaps_df)} staffing requirement gaps detected. Review the table below before publishing this roster.")

st.subheader("Monthly Staffing Plan")
st.caption("Edit the required headcount by date and start time. These values are the recurring planning requirements used for the next draft.")
season_level = st.selectbox("Seasonal demand", ["Low", "Normal", "High"], index=1, key="season_level")
holiday_dates_text = st.text_input("Public holidays (YYYY-MM-DD, comma-separated)", key="holiday_dates_text")
holiday_dates = [x.strip() for x in holiday_dates_text.split(",") if x.strip()]
plan_df = staffing_plan_table(st.session_state.get("schedule_month", date.today().replace(day=1)))
if st.button("Suggest Staffing Plan", key="suggest_staffing_plan"):
    multiplier = {"Low": 0, "Normal": 0, "High": 1}[season_level]
    for idx, row in plan_df.iterrows():
        current_date = str(row["Date"]) or str(plan_df.iloc[idx - 1]["Date"])
        weekend_or_holiday = datetime.strptime(current_date, "%Y-%m-%d").weekday() >= 5 or current_date in holiday_dates
        if weekend_or_holiday:
            for dept in ("FO", "FB", "HK"):
                if dept in plan_df.columns:
                    plan_df.at[idx, dept] = int(plan_df.at[idx, dept]) + 1 + multiplier
        elif multiplier and "FO" in plan_df.columns:
            plan_df.at[idx, "FO"] = int(plan_df.at[idx, "FO"]) + multiplier
    st.session_state["suggested_staffing_plan"] = plan_df
    st.info("Suggested staffing plan created from base requirements, seasonal demand, weekends, and public holidays. Review it before applying.")
if "suggested_staffing_plan" in st.session_state:
    plan_df = st.session_state["suggested_staffing_plan"]
if st.button("Refresh staffing plan table", key="refresh_staffing_plan"):
    st.session_state.pop("staffing_plan_editor", None)
    st.session_state.pop("suggested_staffing_plan", None)
    st.rerun()
edited_plan = st.data_editor(
    plan_df,
    width="content",
    hide_index=True,
    disabled=["Date", "Hour"],
    column_config={"Date": st.column_config.TextColumn("Date", width="medium"), "Hour": st.column_config.TextColumn("Hour", width="small")},
    key="staffing_plan_editor",
)
if st.button("Apply Staffing Plan", key="apply_staffing_plan"):
    for _, plan_row in edited_plan.iterrows():
        hour_key = str(plan_row["Hour"]).split(":")[0].zfill(2)
        for dept in STAFFING_NEEDS:
            if dept in plan_row:
                STAFFING_NEEDS[dept][hour_key] = max(0, int(plan_row[dept]))
    st.success("Staffing plan updated. Generate a new draft to apply these requirements.")

# AI assistance intentionally omitted: roster generation and validation are rule-based.

# 3. Daily Department Coverage
st.header("Who is in charge?")
st.session_state.setdefault(
    "working_date",
    st.session_state.get("schedule_month", date.today().replace(day=1)),
)
working_date = st.date_input(
    "Working date",
    format="YYYY-MM-DD",
    key="working_date",
)
working_month = working_date.replace(day=1)
schedule_is_monthly = all(len(str(k)) == 10 and str(k)[4] == "-" for k in st.session_state["schedule"].keys())
if working_month.replace(day=1) != st.session_state.get("schedule_month") or not schedule_is_monthly:
    st.session_state["schedule_month"] = working_month.replace(day=1)
    st.session_state["schedule"] = generate_month_schedule(employees, st.session_state["schedule_month"], current_preferences, STAFFING_NEEDS)
coverage_day = working_date.strftime("%Y-%m-%d")
st.caption(f"Working date: {coverage_day} · Roster month: {working_month.strftime('%B %Y')}.")

view_mode = st.segmented_control(
    "View mode",
    ["Full day", "Day shift (06-18)", "Night shift (18-06)"],
    default="Full day",
)

if view_mode == "Full day":
    view_hours = hours
elif view_mode == "Day shift (06-18)":
    view_hours = [h for h in hours if 6 <= int(h) <= 18]
else:
    view_hours = [h for h in hours if int(h) >= 18 or int(h) < 6]

coverage_df = schedule_to_department_time_view(st.session_state["schedule"], employees, coverage_day, view_hours)
if st.button("Refresh coverage table", key="refresh_coverage_table"):
    st.rerun()
st.dataframe(coverage_df, width="stretch", hide_index=True)

# 3b. Employee Work Schedule
st.header("Employee Work Schedule")
roster_df = schedule_to_month_employee_view(st.session_state["schedule"], employees, working_month, view_hours)
filter_col1, filter_col2 = st.columns(2)
with filter_col1:
    dept_filter = st.multiselect("Departments", sorted(roster_df["Dept"].unique()), default=sorted(roster_df["Dept"].unique()))
with filter_col2:
    employee_search = st.text_input("Search employee")
filtered = roster_df[roster_df["Dept"].isin(dept_filter)]
if employee_search:
    filtered = filtered[filtered["Employee"].str.contains(employee_search, case=False, na=False)]
st.caption("Edit a date cell to OFF, then click Apply Day-Off Changes.")
if st.button("Refresh employee schedule table", key="refresh_employee_schedule"):
    st.session_state.pop("monthly_roster_editor", None)
    st.rerun()
editable_roster = st.data_editor(
    filtered,
    width="content",
    hide_index=True,
    disabled=["Employee", "Dept"],
    key="monthly_roster_editor",
)
if st.button("Apply Day-Off Changes", key="apply_dayoff_table"):
    editor_state = st.session_state.get("monthly_roster_editor", {})
    changed = apply_monthly_schedule_edits(
        st.session_state["schedule"],
        employees,
        roster_df,
        filtered,
        editor_state.get("edited_rows", {}),
        working_month,
        current_preferences,
    )
    save_preferences(current_preferences)
    st.success(f"Applied {changed} schedule change(s). Hours Summary and coverage are now updated.")
    st.rerun()

if False:
    _old_main_preference_editor = """
# Step 3: manager preference adjustment, placed after the employee schedule.
st.header("Step 3 — Adjust Staff Day-Off Preferences")
st.caption("Upload employees first, generate the draft, then adjust preferences here and regenerate when ready.")
selected_pref_emp_main = st.selectbox(
    "Employee",
    employees,
    format_func=lambda e: f"{e['id']}: {e['name']} ({e['role']})",
    key="main_pref_employee",
)
emp_prefs_main = current_preferences.get(selected_pref_emp_main["id"], {})
monthly_pref_dates_main = sorted(d for d, v in emp_prefs_main.items() if len(str(d)) == 10 and v is None)
preference_options_main = days + monthly_pref_dates_main
selected_days_off_main = st.multiselect(
    "Whole days off / preferred dates",
    preference_options_main,
    default=[d for d, v in emp_prefs_main.items() if v is None and d in preference_options_main],
    key="main_pref_days",
)
pref_col1, pref_col2 = st.columns(2)
with pref_col1:
    if st.button("Save Preferences", key="save_preferences_main"):
        new_prefs = {k: v for k, v in emp_prefs_main.items() if k not in preference_options_main}
        for pref_day in preference_options_main:
            if pref_day in selected_days_off_main:
                new_prefs[pref_day] = None
        current_preferences[selected_pref_emp_main["id"]] = new_prefs
        save_preferences(current_preferences)
        st.success(f"Saved preferences for {selected_pref_emp_main['name']}. Click Generate Rule-Based Schedule to update the draft.")
with pref_col2:
    if st.button("Reset Preferences", key="reset_preferences_main"):
        if os.path.exists(PREFERENCES_FILE):
            os.remove(PREFERENCES_FILE)
        current_preferences = DEFAULT_PREFERENCES.copy()
        st.success("Preferences reset to defaults.")

"""
# 3c. Monthly Hours Summary
st.header("Hours Summary")
hours_df = schedule_to_month_hours_summary(st.session_state["schedule"], employees, working_month)
if st.button("Refresh hours summary table", key="refresh_hours_summary"):
    st.rerun()
st.dataframe(style_month_table(hours_df, working_month), width="content", hide_index=True)

if False:
    _old_manual_adjustment = """
# 4. Manual Adjustment
st.header("Manual Adjustments")
st.caption("You can tick employees directly in this table for the selected date and hour.")
edit_day_options = list(st.session_state["schedule"].keys())
edit_day = st.selectbox("Date to edit", edit_day_options, key="edit_day")
edit_hour = st.selectbox("Hour to edit", hours, format_func=lambda h: f"{h}:00", key="edit_hour")
assigned_now = set(st.session_state["schedule"].get(edit_day, {}).get(edit_hour, []))
edit_df = pd.DataFrame([
    {"Employee": e["name"], "Dept": e["role"], "Assigned": e["id"] in assigned_now}
    for e in employees
])
edited_df = st.data_editor(edit_df, width="content", hide_index=True, disabled=["Employee", "Dept"], key="assignment_editor")
if st.button("Apply table changes", key="apply_table_changes"):
    name_to_id = {e["name"]: e["id"] for e in employees}
    st.session_state["schedule"][edit_day][edit_hour] = [name_to_id[name] for name, assigned in zip(edited_df["Employee"], edited_df["Assigned"]) if assigned]
    st.success("Table changes applied to the draft roster.")

col1, col2, col3 = st.columns(3)

with col1:
    selected_day = st.selectbox("Date", edit_day_options, key="manual_date")
with col2:
    selected_hour = st.selectbox("Hour", hours, format_func=lambda h: f"{h}:00")
with col3:
    selected_emp = st.selectbox(
        "Employee to add/remove",
        [f"{e['id']}: {e['name']} ({e['role']})" for e in employees],
    )

if st.button("Toggle Assignment"):
    emp_id_text = selected_emp.split(":", 1)[0].strip()
    matching_employee = next((e for e in employees if str(e["id"]) == emp_id_text), None)
    if matching_employee is None:
        st.error("Could not find the selected employee ID.")
        st.stop()
    emp_id = matching_employee["id"]
    current_list = st.session_state["schedule"][selected_day][selected_hour]

    if emp_id in current_list:
        current_list.remove(emp_id)
        st.toast(f"Removed {selected_emp.split(': ')[1].split(' (')[0]} from {selected_day} {selected_hour}:00")
    else:
        trial = {d: {h: list(ids) for h, ids in slots.items()} for d, slots in st.session_state["schedule"].items()}
        trial[selected_day][selected_hour].append(emp_id)
        trial_errors = check_rules(trial, employees, current_preferences, hours)
        if any(selected_day in e and selected_hour in e for e in trial_errors):
            st.error("This change creates a conflict for the selected employee/day/hour. Review the conflict list before forcing it.")
            st.stop()
        current_list.append(emp_id)
        st.toast(f"Added {selected_emp.split(': ')[1].split(' (')[0]} to {selected_day} {selected_hour}:00")

    st.rerun()

"""
# 5. Export
st.header("Export Roster")
export_format = st.radio("Choose format:", ["Excel", "CSV"], horizontal=True)

roster_export = schedule_to_month_employee_view(st.session_state["schedule"], employees, working_month, hours)
hours_export = schedule_to_month_hours_summary(st.session_state["schedule"], employees, working_month)

if export_format == "Excel":
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        roster_export.to_excel(writer, index=False, sheet_name="Roster")
        hours_export.to_excel(writer, index=False, sheet_name="Hours Summary")
        plan_df.to_excel(writer, index=False, sheet_name="Staffing Plan")
        blue = "0E5294"
        white = "FFFFFF"
        week_border = Side(style="medium", color=blue)
        for sheet_name in ("Roster", "Hours Summary", "Staffing Plan"):
            ws = writer.book[sheet_name]
            for cell in ws[1]:
                cell.font = Font(bold=True, color=white)
                cell.fill = PatternFill("solid", fgColor=blue)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            for col in range(3, ws.max_column + 1):
                header = ws.cell(1, col).value
                if isinstance(header, str) and len(header) >= 6:
                    try:
                        current = datetime.strptime(f"{header} {working_month.year}", "%d %b %Y")
                        if current.weekday() == 6:
                            for row in range(1, ws.max_row + 1):
                                ws.cell(row, col).border = Border(left=week_border)
                    except ValueError:
                        pass
            for column_cells in ws.columns:
                letter = column_cells[0].column_letter
                ws.column_dimensions[letter].width = min(24, max(12, max(len(str(c.value or "")) for c in column_cells) + 2))
            ws.freeze_panes = "C2"
        # Merge the repeated date cells in the planning sheet for readability.
        plan_ws = writer.book["Staffing Plan"]
        start_row = 2
        for row in range(3, plan_ws.max_row + 2):
            if row == plan_ws.max_row + 1 or plan_ws.cell(row, 1).value:
                end_row = row - 1
                if end_row > start_row:
                    plan_ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                start_row = row
    st.download_button(
        label="Download Excel",
        data=buffer.getvalue(),
        file_name="roster.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.download_button(
        label="Download Roster CSV",
        data=roster_export.to_csv(index=False).encode("utf-8"),
        file_name="roster.csv",
        mime="text/csv",
    )

st.caption(f"Last refreshed: {datetime.now().strftime('%Y-%m-%d %H:%M')} · Review staffing gaps before publishing.")
